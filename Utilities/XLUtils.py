import openpyxl


class Excel_methods:

    @staticmethod
    def read_data_from_excel(file_path,sheet_name,row_no,column_no):
        excel_file=openpyxl.load_workbook(file_path)
        sheet=excel_file[sheet_name]
        return sheet.cell(row_no,column_no).value

    @staticmethod
    def write_data_to_excel(file_path, sheet_name, row_no, column_no,data):
        excel_file = openpyxl.load_workbook(file_path)
        sheet = excel_file[sheet_name]
        sheet.cell(row_no, column_no).value=data
        excel_file.save(file_path)
        excel_file.close()

    @staticmethod
    def get_total_rowcount(file_path, sheet_name):
        excel_file = openpyxl.load_workbook(file_path)
        sheet = excel_file[sheet_name]
        return sheet.max_row
